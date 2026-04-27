#!/usr/bin/env python3
"""
ETL: Parse MOHW food safety xlsx files → SQL for dashboard DB.

Sources (docs/assets/):
  - 10521-05-01食品中毒案件病因物質分類統計.xlsx  → food_poisoning_national
  - 10521-05-03食品中毒案件攝食場所分類統計.xlsx  → food_poisoning_location
  - 10521-01-03食品衛生管理工作－按縣市別分1150331.xlsx → food_inspection_by_city

Output: /tmp/mohw_food_stats.sql → inject into postgres-data (dashboard DB)

Run from repo root:
  python scripts/generate_mohw_food_stats_sql.py
"""
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).parent.parent
ASSETS    = REPO_ROOT / "docs" / "assets"

F_POISONING_CAUSE    = ASSETS / "10521-05-01食品中毒案件病因物質分類統計.xlsx"
F_POISONING_LOCATION = ASSETS / "10521-05-03食品中毒案件攝食場所分類統計.xlsx"
F_BY_CITY            = ASSETS / "10521-01-03食品衛生管理工作－按縣市別分1150331.xlsx"

OUT_SQL = Path("/tmp/mohw_food_stats.sql")


# ── helpers ──────────────────────────────────────────────────────────────────

def to_int(v):
    if v is None:
        return 0
    try:
        return int(str(v).replace(",", "").strip())
    except (ValueError, AttributeError):
        return 0

def to_float(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, AttributeError):
        return 0.0

def roc_to_ad(roc_year_str):
    """Convert ROC year like '113年' or '113' to AD year int."""
    s = str(roc_year_str).strip()
    m = re.search(r"(\d+)", s)
    if not m:
        return None
    roc = int(m.group(1))
    if roc < 50:   # likely already AD year fragment
        return None
    return roc + 1911


# ── 1. food_poisoning_national (from 2.1.8 歷年資料) ─────────────────────────

def parse_poisoning_national():
    wb = openpyxl.load_workbook(F_POISONING_CAUSE, data_only=True)
    ws = wb["歷年資料"]
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Row format: 年別 | AD year | 件數 | 患者數 | 死亡數
        # e.g. '113年' | 2024 | 1750 | 9045 | 10
        if row[0] is None:
            continue
        year = roc_to_ad(row[0])
        if year is None:
            # try second column (some rows have AD year there)
            if row[1]:
                try:
                    year = int(row[1])
                    if year < 1990 or year > 2030:
                        continue
                except:
                    continue
            else:
                continue
        cases    = to_int(row[2])
        patients = to_int(row[3])
        deaths   = to_int(row[4])
        if cases == 0 and patients == 0:
            continue
        rows.append((year, cases, patients, deaths))
    print(f"  food_poisoning_national: {len(rows)} rows ({rows[0][0]}–{rows[-1][0]})")
    return rows


# ── 2. food_poisoning_location (from 2.1.10 each year sheet) ─────────────────

LOCATION_MAP = {
    "自宅":       "自宅",
    "供膳之營業場所": "餐飲業",
    "學校":       "學校",
    "辦公場所":   "辦公場所",
    "醫療場所":   "醫療場所",
    "運輸工具":   "運輸工具",
    "部隊":       "部隊",
    "野外":       "野外",
    "攤販":       "攤販",
    "外燴":       "外燴",
    "監獄":       "監獄",
    "其他":       "其他",
}

def parse_poisoning_location():
    wb = openpyxl.load_workbook(F_POISONING_LOCATION, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        if "歷年" in sheet_name or "說明" in sheet_name:
            continue
        year = roc_to_ad(sheet_name)
        if year is None:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            location_raw = str(row[0]).strip() if row[0] else ""
            # strip leading digits like "1總計" → skip 總計
            location_raw = re.sub(r"^\d+", "", location_raw).strip()
            if not location_raw or location_raw == "總計":
                continue
            # skip footer rows
            if any(kw in location_raw for kw in ["填表", "資料來源", "填表說明", "審核"]):
                continue
            # map to clean name
            location = LOCATION_MAP.get(location_raw, location_raw)
            cases    = to_int(row[2])
            patients = to_int(row[3])
            deaths   = to_int(row[4])
            if cases == 0 and patients == 0:
                continue
            rows.append((year, location, cases, patients, deaths))
    print(f"  food_poisoning_location: {len(rows)} rows across {len(set(r[0] for r in rows))} years")
    return rows


# ── 3. food_inspection_by_city (from 2.1.3 each year sheet) ──────────────────

CITY_MAP = {
    "新  北  市": "新北市",
    "新北市":     "新北市",
    "臺  北  市": "臺北市",
    "臺北市":     "臺北市",
    "台北市":     "臺北市",
    "桃 園 市":   "桃園市",
    "桃園市":     "桃園市",
    "臺 中 市":   "臺中市",
    "臺中市":     "臺中市",
    "臺 南 市":   "臺南市",
    "臺南市":     "臺南市",
    "高  雄  市": "高雄市",
    "高雄市":     "高雄市",
    "宜  蘭  縣": "宜蘭縣",
    "宜蘭縣":     "宜蘭縣",
    "新 竹 縣":   "新竹縣",
    "新竹縣":     "新竹縣",
    "苗 栗 縣":   "苗栗縣",
    "苗栗縣":     "苗栗縣",
    "彰 化 縣":   "彰化縣",
    "彰化縣":     "彰化縣",
    "南 投 縣":   "南投縣",
    "南投縣":     "南投縣",
    "雲 林 縣":   "雲林縣",
    "雲林縣":     "雲林縣",
    "嘉 義 縣":   "嘉義縣",
    "嘉義縣":     "嘉義縣",
    "屏 東 縣":   "屏東縣",
    "屏東縣":     "屏東縣",
    "花 蓮 縣":   "花蓮縣",
    "花蓮縣":     "花蓮縣",
    "臺 東 縣":   "臺東縣",
    "臺東縣":     "臺東縣",
    "澎 湖 縣":   "澎湖縣",
    "澎湖縣":     "澎湖縣",
    "金 門 縣":   "金門縣",
    "金門縣":     "金門縣",
    "連 江 縣":   "連江縣",
    "連江縣":     "連江縣",
    "新  竹  市": "新竹市",
    "新竹市":     "新竹市",
    "嘉  義  市": "嘉義市",
    "嘉義市":     "嘉義市",
    "基  隆  市": "基隆市",
    "基隆市":     "基隆市",
}

def normalize_city(raw):
    raw = str(raw).strip()
    # try exact match first
    if raw in CITY_MAP:
        return CITY_MAP[raw]
    # collapse all whitespace variants and try again
    collapsed = re.sub(r"\s+", "", raw).strip()  # remove all spaces
    for key, val in CITY_MAP.items():
        if re.sub(r"\s+", "", key) == collapsed:
            return val
    return None

def parse_inspection_by_city():
    wb = openpyxl.load_workbook(F_BY_CITY, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        if "歷年" in sheet_name or "說明" in sheet_name:
            continue
        year = roc_to_ad(sheet_name)
        if year is None:
            continue
        ws = wb[sheet_name]
        current_city = None
        insp_count   = None
        for row in ws.iter_rows(values_only=True):
            col0 = str(row[0]).strip() if row[0] else ""
            col2 = str(row[2]).strip() if row[2] else ""

            # Detect city row (col0 has city name, col2 = '查驗件數')
            city = normalize_city(col0)
            if city:
                current_city = city
                if col2 == "查驗件數":
                    insp_count = to_int(row[4])
                continue

            # 不符規定件數 follows immediately after city 查驗件數 row
            if current_city and col2 == "不符規定件數" and insp_count is not None:
                violation_count = to_int(row[4])
                violation_rate  = round(violation_count * 100 / insp_count, 4) if insp_count else 0
                rows.append((year, current_city, insp_count, violation_count, violation_rate))
                current_city = None
                insp_count   = None
    print(f"  food_inspection_by_city: {len(rows)} rows, years {sorted(set(r[0] for r in rows))}")
    return rows


# ── SQL generation ────────────────────────────────────────────────────────────

def esc(s):
    return str(s).replace("'", "''")

def generate_sql(poisoning_nat, poisoning_loc, inspection_city):
    lines = [
        "-- MOHW food safety national statistics",
        "-- Generated by scripts/generate_mohw_food_stats_sql.py",
        "",
        "-- 1. food_poisoning_national",
        "DROP TABLE IF EXISTS food_poisoning_national;",
        "CREATE TABLE food_poisoning_national (",
        "  year           INT PRIMARY KEY,",
        "  cases          INT NOT NULL DEFAULT 0,",
        "  patients       INT NOT NULL DEFAULT 0,",
        "  deaths         INT NOT NULL DEFAULT 0",
        ");",
    ]
    for year, cases, patients, deaths in poisoning_nat:
        lines.append(
            f"INSERT INTO food_poisoning_national VALUES ({year},{cases},{patients},{deaths});"
        )

    lines += [
        "",
        "-- 2. food_poisoning_location",
        "DROP TABLE IF EXISTS food_poisoning_location;",
        "CREATE TABLE food_poisoning_location (",
        "  year           INT NOT NULL,",
        "  location       TEXT NOT NULL,",
        "  cases          INT NOT NULL DEFAULT 0,",
        "  patients       INT NOT NULL DEFAULT 0,",
        "  deaths         INT NOT NULL DEFAULT 0",
        ");",
    ]
    for year, location, cases, patients, deaths in poisoning_loc:
        lines.append(
            f"INSERT INTO food_poisoning_location VALUES ({year},'{esc(location)}',{cases},{patients},{deaths});"
        )

    lines += [
        "",
        "-- 3. food_inspection_by_city",
        "DROP TABLE IF EXISTS food_inspection_by_city;",
        "CREATE TABLE food_inspection_by_city (",
        "  year              INT NOT NULL,",
        "  city              TEXT NOT NULL,",
        "  inspection_count  INT NOT NULL DEFAULT 0,",
        "  violation_count   INT NOT NULL DEFAULT 0,",
        "  violation_rate    NUMERIC(8,4) NOT NULL DEFAULT 0",
        ");",
    ]
    for year, city, insp, viol, rate in inspection_city:
        lines.append(
            f"INSERT INTO food_inspection_by_city VALUES ({year},'{esc(city)}',{insp},{viol},{rate});"
        )

    lines.append("")
    return "\n".join(lines)


# ── main ──────────────────────────────────────────────────────────────────────

def main():
    print("Parsing xlsx files...")
    poisoning_nat  = parse_poisoning_national()
    poisoning_loc  = parse_poisoning_location()
    inspection_city = parse_inspection_by_city()

    sql = generate_sql(poisoning_nat, poisoning_loc, inspection_city)
    OUT_SQL.write_text(sql, encoding="utf-8")
    print(f"\nSQL written to {OUT_SQL}")
    print(f"  food_poisoning_national:  {len(poisoning_nat)} rows")
    print(f"  food_poisoning_location:  {len(poisoning_loc)} rows")
    print(f"  food_inspection_by_city:  {len(inspection_city)} rows")

if __name__ == "__main__":
    main()
