#!/usr/bin/env python3
"""
ETL: Parse MOHW 2.1.3 xlsx (food inspection by city) → food_type_violations SQL.

Extracts per-city violation counts broken down by major food category.
Covers both 臺北市 and 新北市, enabling genuine dual-city 1014 chart.

Source: docs/assets/10521-01-03食品衛生管理工作－按縣市別分1150331.xlsx
Output: /tmp/food_type_violations.sql
"""
import re
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).parent.parent
F_BY_CITY = REPO_ROOT / "docs" / "assets" / "10521-01-03食品衛生管理工作－按縣市別分1150331.xlsx"
OUT_SQL = Path("/tmp/food_type_violations.sql")

TARGET_CITIES = {"臺北市", "新北市"}

# Column index → (category_label, section)
# Section 1: cols 5-35, Section 2: cols 40-70
# Each tuple: list of column indices that belong to the category
CATEGORY_COLS = {
    "乳品類":     list(range(5, 11)),           # dairy
    "肉品類":     list(range(11, 14)),           # meat
    "蛋品類":     list(range(14, 17)),           # eggs
    "水產類":     list(range(17, 20)),           # seafood
    "穀豆烘焙":   list(range(20, 32)),           # grains, beans, rice, noodles, bread, candy
    "蔬果類":     list(range(32, 36)) + list(range(40, 44)),  # fruits/vegetables (split)
    "飲料及水":   list(range(49, 52)),           # beverages and water
    "食用油脂":   list(range(52, 55)),           # edible oils
    "調味品":     list(range(59, 63)),           # cleaners + condiments + soy sauce
    "健康食品":   [63, 64],                      # health claim + health food
    "複合調理":   list(range(65, 68)),           # composite processed food
    "其他":       list(range(44, 49)) + list(range(55, 59)) + list(range(68, 72)),  # rest
}

CITY_MAP = {
    "新北市": "新北市",
    "臺北市": "臺北市",
}

def normalize_city(raw):
    s = re.sub(r"\s+", "", str(raw).strip())
    for key in CITY_MAP:
        if re.sub(r"\s+", "", key) == s:
            return CITY_MAP[key]
    return None

def to_int(v):
    if v is None:
        return 0
    try:
        return int(str(v).replace(",", "").strip())
    except (ValueError, AttributeError):
        return 0

def roc_to_ad(s):
    m = re.search(r"(\d+)", str(s))
    if not m:
        return None
    roc = int(m.group(1))
    return roc + 1911 if roc >= 50 else None

def sum_cols(row, cols):
    return sum(to_int(row[c]) if c < len(row) else 0 for c in cols)

def parse_sheet(ws, year):
    rows_data = []
    all_rows = list(ws.iter_rows(values_only=True))
    current_city = None

    for i, row in enumerate(all_rows):
        col0 = str(row[0]).strip() if row[0] else ""
        col2 = str(row[2]).strip() if row[2] else ""

        # Detect city row
        city = normalize_city(col0) if col0 else None
        if city and city in TARGET_CITIES:
            current_city = city
            continue

        # Detect violation row (follows city inspection row)
        if current_city and ("不符規定" in col2 or (col2 == "" and row[0] is None and i > 0)):
            # Look for the actual violation indicator
            if "不符規定" not in col2:
                continue
            viol_by_cat = {}
            for cat, cols in CATEGORY_COLS.items():
                viol_by_cat[cat] = sum_cols(row, cols)
            rows_data.append((year, current_city, viol_by_cat))
            current_city = None  # reset after consuming

    return rows_data

def main():
    wb = openpyxl.load_workbook(F_BY_CITY, data_only=True)
    all_rows = []

    for sheet_name in wb.sheetnames:
        if "歷年" in sheet_name or "說明" in sheet_name:
            continue
        year = roc_to_ad(sheet_name)
        if year is None:
            continue
        ws = wb[sheet_name]
        rows = parse_sheet(ws, year)
        all_rows.extend(rows)

    years_found = sorted(set(r[0] for r in all_rows))
    cities_found = sorted(set(r[1] for r in all_rows))
    print(f"Parsed {len(all_rows)} city-year rows: years {years_found[0]}-{years_found[-1]}, cities {cities_found}")

    # Generate SQL
    lines = [
        "-- Food violation counts by food category per city",
        "-- Source: MOHW 2.1.3 (10521-01-03)",
        "",
        "DROP TABLE IF EXISTS food_type_violations;",
        "CREATE TABLE food_type_violations (",
        "  year            INT NOT NULL,",
        "  city            TEXT NOT NULL,",
        "  food_type       TEXT NOT NULL,",
        "  violation_count INT NOT NULL DEFAULT 0",
        ");",
    ]

    for year, city, viol_by_cat in all_rows:
        for cat, count in viol_by_cat.items():
            city_esc = city.replace("'", "''")
            cat_esc = cat.replace("'", "''")
            lines.append(
                f"INSERT INTO food_type_violations VALUES ({year},'{city_esc}','{cat_esc}',{count});"
            )

    lines.append("")
    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    print(f"SQL written to {OUT_SQL}")

    # Print sample
    sample_year = years_found[-1]
    print(f"\nSample for {sample_year}:")
    for year, city, viol_by_cat in all_rows:
        if year == sample_year:
            print(f"  {city}: {viol_by_cat}")

if __name__ == "__main__":
    main()
