#!/usr/bin/env python3
"""
Load MOHW national food-poisoning statistics by cause (all available years).

Source (xlsx only — NO HTTP):
  docs/assets/10521-05-01食品中毒案件病因物質分類統計.xlsx

Reads per-year sheets (e.g. '113年 ', '112年', ...).
Each sheet has rows of the form:
  col[0]=main_category (e.g. '細菌', '天然毒'), col[1]=sub_cause (specific name),
  col[2]=cases, col[3]=persons, col[4]=deaths

Strategy:
  - Skip header rows (rows 0-4), summary aggregates, subtotal rows, footer rows.
  - For rows with a specific sub-cause in col[1]: use col[1] as cause label.
  - For standalone rows where col[0] has a category but col[1] is None:
    use col[0] as cause label (e.g. '其他病因物質').
  - Only store rows where cases > 0 OR persons > 0.

Writes food_poisoning_cause (year, cause, cases, persons).
Adapted from scripts/generate_mohw_food_stats_sql.py (parent worktree).
"""
import re
import sys
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _db import db_kwargs  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parents[3]
XLSX_PATH = REPO_ROOT / "docs" / "assets" / "10521-05-01食品中毒案件病因物質分類統計.xlsx"

# Rows to skip at the top-level (aggregate / footer rows)
SKIP_COL0_PREFIXES = (
    "總計", "病因物質判明合計", "病因物質不明合計",
    "填表", "資料來源", "附", "說明", "食品中毒",
    "Food", "中華民國", "單位", "病  因",
)

# col[1] values that are subtotals rather than actual causes
SKIP_COL1_PREFIXES = ("小計",)

# When the sub-cause is a generic label like '其他', qualify it with the parent
# category to avoid (year, cause) uniqueness conflicts across categories.
GENERIC_CAUSES = {"其他"}

INSERT_SQL = "INSERT INTO food_poisoning_cause (year, cause, cases, persons) VALUES %s"


def to_int(v):
    try:
        return int(str(v or "0").replace(",", "").strip())
    except (ValueError, AttributeError):
        return 0


def roc_to_ad(s):
    m = re.search(r"(\d+)", str(s))
    if not m:
        return None
    roc = int(m.group(1))
    return roc + 1911 if roc >= 50 else None


def clean_label(s):
    """Strip trailing footnote digits (e.g. '小計1' → skip, '其他3' → '其他')."""
    return re.sub(r"\d+$", "", str(s or "").strip()).strip()


def parse_sheet(ws, year):
    rows = []
    # Track parent category to qualify generic sub-cause labels (e.g. '其他').
    # Category rows have col[0]=category_name AND col[1] starting with '小計'.
    current_category = None

    for row in ws.iter_rows(values_only=True):
        c0 = str(row[0]).strip() if row[0] is not None else ""
        c1 = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        # Skip header, aggregates, footers identified by col[0]
        if c0 and any(c0.startswith(p) for p in SKIP_COL0_PREFIXES):
            continue

        # Main category row: col[0] has category, col[1] has '小計...'
        # Update current_category but don't emit a row (subtotals are skipped).
        if c0 and c1 and any(c1.startswith(p) for p in SKIP_COL1_PREFIXES):
            current_category = clean_label(c0)
            continue

        # Determine cause label
        if c1:
            sub = clean_label(c1)
            # Qualify generic names (e.g. '其他') with parent category
            if sub in GENERIC_CAUSES and current_category:
                cause = f"{current_category}-{sub}"
            else:
                cause = sub
        elif c0:
            # Standalone category row (e.g. '其他病因物質', '病毒' sub-rows w/o subtotal)
            cause = clean_label(c0)
            # Also update current_category
            current_category = cause
        else:
            continue

        cases   = to_int(row[2]) if len(row) > 2 else 0
        persons = to_int(row[3]) if len(row) > 3 else 0

        if cases == 0 and persons == 0:
            continue

        rows.append((year, cause, cases, persons))
    return rows


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    out = []

    for sheet_name in wb.sheetnames:
        if "歷年" in sheet_name or "說明" in sheet_name:
            continue
        year = roc_to_ad(sheet_name)
        if year is None:
            continue
        ws = wb[sheet_name]
        rows = parse_sheet(ws, year)
        out.extend(rows)

    if not out:
        print("No rows parsed — check xlsx structure.")
        return

    years = sorted(set(r[0] for r in out))
    print(f"Parsed {len(out)} rows across {len(years)} years ({years[0]}–{years[-1]})")

    with psycopg2.connect(**db_kwargs()) as conn, conn.cursor() as cur:
        cur.execute("TRUNCATE food_poisoning_cause RESTART IDENTITY")
        execute_values(cur, INSERT_SQL, out)
        conn.commit()

    print(f"food_poisoning_cause: {len(out)} rows inserted")


if __name__ == "__main__":
    main()
