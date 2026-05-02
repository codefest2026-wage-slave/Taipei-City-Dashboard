#!/usr/bin/env python3
"""
Load MOHW dual-city food inspection statistics for all available years.

Source (xlsx only — NO HTTP):
  docs/assets/10521-01-03食品衛生管理工作－按縣市別分1150331.xlsx

Writes two tables:
  - food_inspection_by_city   (year, city, venue='合計', inspections,
                               noncompliance, poisoning_cases=NULL,
                               ntpc_violation_rate)
  - food_type_violations      (year, city, category, count)

Structure of each per-year sheet:
  - Each city occupies two consecutive rows:
      Row A: col[0]=city_name, col[2]='查驗件數',    col[4]=total_inspections
      Row B: col[0]=None,      col[2]='不符規定件數', col[4]=total_nc,
             cols[5-70]=per-category NC counts

Column-index maps adapted from:
  scripts/generate_mohw_food_stats_sql.py
  scripts/generate_food_type_violations_sql.py  (parent worktree, validated)
"""
import re
import sys
from pathlib import Path

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _db import db_kwargs  # noqa: E402

REPO_ROOT = Path(__file__).resolve().parents[3]
XLSX_PATH = REPO_ROOT / "docs" / "assets" / "10521-01-03食品衛生管理工作－按縣市別分1150331.xlsx"

TARGET_CITIES = {"臺北市", "新北市"}

# Category → list of column indices in the NC row.
# Validated: sum(cols) == col[4] (total NC) for both cities in 114年.
CATEGORY_COLS = {
    "乳品類":     list(range(5, 11)),
    "肉品類":     list(range(11, 14)),
    "蛋品類":     list(range(14, 17)),
    "水產類":     list(range(17, 20)),
    "穀豆烘焙":   list(range(20, 32)),
    "蔬果類":     list(range(32, 36)) + list(range(40, 44)),
    "飲料及水":   list(range(49, 52)),
    "食用油脂":   list(range(52, 55)),
    "調味品":     list(range(59, 63)),
    "健康食品":   [63, 64],
    "複合調理":   list(range(65, 68)),
    "其他":       list(range(44, 49)) + list(range(55, 59)) + list(range(68, 72)),
}

CITY_MAP = {
    "新北市": "新北市",
    "臺北市": "臺北市",
    "台北市": "臺北市",
}


def to_int(v):
    try:
        return int(str(v or "0").replace(",", "").strip())
    except (ValueError, AttributeError):
        return 0


def roc_to_ad(s):
    m = re.search(r"(\d+)", str(s))
    if not m:
        return None
    roc = int(m.group(1))
    return roc + 1911 if roc >= 50 else None


def normalize_city(raw):
    s = re.sub(r"\s+", "", str(raw or "").strip())
    for key, val in CITY_MAP.items():
        if re.sub(r"\s+", "", key) == s:
            return val
    return None


def parse_sheet(ws, year):
    """Return (inspection_rows, violation_rows) for one year sheet."""
    inspection = []
    violations = []

    current_city = None
    insp_count = None

    for row in ws.iter_rows(values_only=True):
        col0 = str(row[0]).strip() if row[0] else ""
        col2 = str(row[2]).strip() if row[2] else ""

        # City row: col[0] has city name, col[2]='查驗件數'
        city = normalize_city(col0)
        if city and city in TARGET_CITIES and col2 == "查驗件數":
            current_city = city
            insp_count = to_int(row[4])
            continue

        # NC row: col[0] is None/empty, col[2]='不符規定件數'
        if current_city and col2 == "不符規定件數" and insp_count is not None:
            nc_count = to_int(row[4])
            rate = round(nc_count * 100.0 / insp_count, 2) if insp_count else None
            # poisoning_cases: not available in this xlsx (stored as NULL)
            inspection.append((year, current_city, "合計", insp_count, nc_count, None, rate))

            # Per-category violation counts (only store > 0)
            for cat, cols in CATEGORY_COLS.items():
                total = sum(to_int(row[c]) for c in cols if c < len(row))
                if total > 0:
                    violations.append((year, current_city, cat, total))

            current_city = None
            insp_count = None

    return inspection, violations


def parse_xlsx():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    all_insp = []
    all_viol = []

    for sheet_name in wb.sheetnames:
        if "歷年" in sheet_name or "說明" in sheet_name:
            continue
        year = roc_to_ad(sheet_name)
        if year is None:
            continue
        ws = wb[sheet_name]
        insp, viol = parse_sheet(ws, year)
        all_insp.extend(insp)
        all_viol.extend(viol)

    return all_insp, all_viol


def main():
    insp_rows, viol_rows = parse_xlsx()

    years = sorted(set(r[0] for r in insp_rows))
    cities = sorted(set(r[1] for r in insp_rows))
    print(f"Parsed: {len(insp_rows)} inspection rows, {len(viol_rows)} violation rows")
    print(f"  Years: {years[0]}–{years[-1]}, Cities: {cities}")

    with psycopg2.connect(**db_kwargs()) as conn, conn.cursor() as cur:
        cur.execute("TRUNCATE food_inspection_by_city RESTART IDENTITY")
        execute_values(
            cur,
            "INSERT INTO food_inspection_by_city "
            "(year, city, venue, inspections, noncompliance, poisoning_cases, ntpc_violation_rate) "
            "VALUES %s",
            insp_rows,
        )
        cur.execute("TRUNCATE food_type_violations RESTART IDENTITY")
        execute_values(
            cur,
            "INSERT INTO food_type_violations (year, city, category, count) VALUES %s",
            viol_rows,
        )
        conn.commit()

    print(f"food_inspection_by_city: {len(insp_rows)} rows inserted")
    print(f"food_type_violations:    {len(viol_rows)} rows inserted")


if __name__ == "__main__":
    main()
